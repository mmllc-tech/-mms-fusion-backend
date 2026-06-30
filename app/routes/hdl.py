"""
HDL Routes
- POST /api/hdl/{project_id}/validate  — validate uploaded client Excel
- POST /api/hdl/{project_id}/generate  — generate HDL files from validated data
- GET  /api/hdl/{project_id}/files     — list generated HDL files
- GET  /api/hdl/{project_id}/sequence  — get the full load sequence
"""
import io, json, os, uuid
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.models.database import get_db, Project, LoadRun, Evidence
from app.services.hdl_generator import (
    build_hdl_file, validate_sheet, get_load_sequence, HDL_OBJECTS
)

router = APIRouter()

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/tmp/mms_fusion_hdl")
os.makedirs(UPLOAD_DIR, exist_ok=True)

SHEET_TO_OBJECT = {
    "Locations": "Location",
    "Departments": "Department",
    "Job Families": "JobFamily",
    "Jobs": "Job",
    "Grades": "Grade",
    "Actions": "Action",
    "Action Reasons": "ActionReason",
    "Reference Data Sets": "ReferenceDataSet",
    "Legislative Data Groups": "LegislativeDataGroup",
    "Legal Entities": "LegalEntity",
    "Business Units": "BusinessUnit",
}

def read_excel_sheets(content: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            vals = [str(v or "").strip() for v in row]
            if any(vals) and not all(v == "" for v in vals):
                header_row = vals
                data_start = i + 1
                break
        if not header_row:
            continue
        import re
        clean_headers = []
        for h in header_row:
            h = str(h or "").strip()
            h = re.sub(r'\s*\[.*?\]', '', h).strip()
            clean_headers.append(h)

        data_rows = []
        for row in rows[data_start:]:
            row_dict = {}
            for col_idx, val in enumerate(row):
                if col_idx < len(clean_headers):
                    header = clean_headers[col_idx]
                    if header and not header.startswith("Comments"):
                        row_dict[header] = str(val or "").strip()
            if any(v for v in row_dict.values()):
                data_rows.append(row_dict)
        if data_rows:
            result[sheet_name] = data_rows
    return result

@router.post("/{project_id}/validate")
async def validate_client_data(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    try:
        sheets = read_excel_sheets(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    results = []
    total_errors = 0
    total_warnings = 0

    for sheet_name, rows in sheets.items():
        validation = validate_sheet(sheet_name, rows)
        results.append(validation)
        total_errors += len(validation["errors"])
        total_warnings += len(validation["warnings"])

    return {
        "file": file.filename,
        "sheets_found": list(sheets.keys()),
        "total_rows": sum(r["total_rows"] for r in results),
        "total_errors": total_errors,
        "total_warnings": total_warnings,
        "can_proceed": total_errors == 0,
        "results": results
    }

@router.post("/{project_id}/generate")
async def generate_hdl_files(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    try:
        sheets = read_excel_sheets(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {str(e)}")

    generated = []
    project_dir = os.path.join(UPLOAD_DIR, project_id)
    os.makedirs(project_dir, exist_ok=True)

    for sheet_name, rows in sorted(
        sheets.items(),
        key=lambda x: SHEET_TO_OBJECT.get(x[0], "ZZ")
    ):
        obj_name = SHEET_TO_OBJECT.get(sheet_name)
        if not obj_name or obj_name not in HDL_OBJECTS:
            continue

        try:
            hdl_content = build_hdl_file(obj_name, rows)
            hdl_filename = f"{obj_name}_{project_id[:8]}.dat"
            hdl_path = os.path.join(project_dir, hdl_filename)
            with open(hdl_path, "w") as f:
                f.write(hdl_content)

            run = LoadRun(
                project_id=project_id,
                object_name=obj_name,
                method=HDL_OBJECTS[obj_name]["method"],
                sequence=HDL_OBJECTS[obj_name]["seq"],
                status="ready",
                hdl_file_name=hdl_filename,
            )
            db.add(run)

            ev = Evidence(
                project_id=project_id,
                evidence_type="hdl_file",
                file_name=hdl_filename,
                file_path=hdl_path,
                notes=f"Generated from {sheet_name} — {len(rows)} records"
            )
            db.add(ev)

            generated.append({
                "object": obj_name,
                "sheet": sheet_name,
                "rows": len(rows),
                "file": hdl_filename,
                "lines": len(hdl_content.split("\n")),
                "preview": "\n".join(hdl_content.split("\n")[:3]) + "\n..."
            })
        except Exception as e:
            generated.append({
                "object": obj_name,
                "sheet": sheet_name,
                "error": str(e)
            })

    db.commit()
    return {
        "generated": len([g for g in generated if "error" not in g]),
        "errors": len([g for g in generated if "error" in g]),
        "files": generated
    }

@router.get("/{project_id}/download/{filename}")
def download_hdl_file(project_id: str, filename: str):
    path = os.path.join(UPLOAD_DIR, project_id, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="File not found")
    with open(path, "r") as f:
        content = f.read()
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{project_id}/files")
def list_hdl_files(project_id: str, db: Session = Depends(get_db)):
    runs = db.query(LoadRun).filter(
        LoadRun.project_id == project_id
    ).order_by(LoadRun.sequence).all()
    return [{
        "id": r.id, "object": r.object_name, "method": r.method,
        "sequence": r.sequence, "status": r.status,
        "file": r.hdl_file_name, "created": r.created_at
    } for r in runs]

@router.get("/{project_id}/sequence")
def get_sequence(project_id: str):
    return {"sequence": get_load_sequence()}
